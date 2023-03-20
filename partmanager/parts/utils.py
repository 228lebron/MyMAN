import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd
from matplotlib.cm import ScalarMappable
from matplotlib.colors import Normalize
import base64
from io import BytesIO
import openpyxl
from .models import Quota, Part, Request, RequestQuotaResult
import plotly.graph_objs as go
from .constants import brand_mapping


def check_ic_series(ic):
    series_list = [i.replace(';', '') for i in pd.read_csv('C:\Manager\PartsManagment\partmanager\parts\IC_Series\AD.csv', header=None)[0].tolist()]
    maybe_series = []
    for series in series_list:
        if ic.startswith(series):
            maybe_series.append(series)
    if len(maybe_series) == 0:
        return None
    else:
        return sorted(maybe_series, key=len, reverse=True)[0]



#def create_quotas_from_xlsx(file):
#    wb = openpyxl.load_workbook(file)
#    ws = wb.active
#    quotas = []
#    for row in ws.iter_rows(values_only=True):
#        part_number, brand, quantity, price, datecode, lead_time, supplier, date = row
#
#        if brand in brand_mapping:
#            brand = brand_mapping[brand]
#
#        try:
#            part = Part.objects.filter(number=part_number, brand=brand).first()
#        except Part.DoesNotExist:
#            part_series = check_ic_series(part_number)
#            part = Part.objects.create(number=part_number, series=part_series, brand=brand)
#        quotas.append(Quota(part=part, quantity=quantity, price=price, datecode=datecode, supplier=supplier,
#                            lead_time=lead_time, date=date))
#    Quota.objects.bulk_create(quotas)

def create_quotas_from_xlsx(file):
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    for row in ws.iter_rows(values_only=True):
        part_number, brand, quantity, price, datecode, lead_time, supplier, date = row

        if brand in brand_mapping:
            brand = brand_mapping[brand]

        try:
            part = Part.objects.filter(number=part_number, brand=brand).first()
        except Part.DoesNotExist:
            part_series = check_ic_series(part_number)
            part = Part.objects.create(number=part_number, series=part_series, brand=brand)

        quota = Quota(part=part, quantity=quantity, price=price, datecode=datecode, supplier=supplier,
                      lead_time=lead_time, date=date)
        quota.save()



#def get_price_data(part):
#    quotas = Quota.objects.filter(part=part)
#    requests = Request.objects.filter(part=part)
#    prices = []
#    suppliers = []
#    dates = []
#    lead_times = []
#    quantity = []
#    for quota in quotas:
#        prices.append(float(quota.price))
#        dates.append(quota.date)
#        suppliers.append(quota.supplier)
#        lead_times.append(quota.lead_time)
#        quantity.append(quota.quantity)
#    return prices, dates, suppliers, lead_times, quantity

# РАБОЧАЯ ФУНКЦИЯ
def get_price_data(part):
    quotas = Quota.objects.filter(part=part)
    requests = Request.objects.filter(part=part)
    prices = []
    suppliers = []
    dates = []
    lead_times = []
    quantity = []
    for quota in sorted(quotas, key=lambda q: q.date):
        prices.append(float(quota.price))
        dates.append(quota.date)
        suppliers.append(quota.supplier)
        lead_times.append(quota.lead_time)
        quantity.append(quota.quantity)
    return prices, dates, suppliers, lead_times, quantity

# словарь для графиков
def get_price_data_d(part):
    quotas = Quota.objects.filter(part=part)
    requests = Request.objects.filter(part=part)
    prices = {}
    dates = {}
    lead_times = {}
    quantity = {}
    for quota in sorted(quotas, key=lambda q: q.date):
        supplier = quota.supplier
        if supplier not in prices:
            prices[supplier] = []
            dates[supplier] = []
            lead_times[supplier] = []
            quantity[supplier] = []
        prices[supplier].append(float(quota.price))
        dates[supplier].append(quota.date)
        lead_times[supplier].append(quota.lead_time)
        quantity[supplier].append(quota.quantity)
    return prices, dates, lead_times, quantity

def get_graph():
    buffer = BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)
    image_png = buffer.getvalue()
    graph = base64.b64encode(image_png)
    graph = graph.decode('utf-8')
    buffer.close()
    return graph

#def get_plot(x, y, title, xlabel, ylabel):
#    plt.switch_backend('AGG')
#    plt.figure(figsize=(9,5))
#    plt.style.use('seaborn-whitegrid')
#    plt.title(title, fontsize=16, fontweight='bold')
#    #plt.plot(x, y)
#    plt.plot(x, y, color='tab:blue', linewidth=2, marker='o', markersize=6, linestyle='--')
#    plt.legend(loc='best', fontsize=12)
#    plt.xlabel(xlabel, fontsize=14)
#    plt.ylabel(ylabel, fontsize=14)
#    plt.xticks(fontsize=8)
#    plt.yticks(fontsize=8)
#    plt.tight_layout()
#    graph = get_graph()
#    return graph

def get_plot(x, y_dict, title, xlabel, ylabel):
    plt.switch_backend('AGG')
    plt.figure(figsize=(9,5))
    plt.style.use('seaborn-whitegrid')
    plt.title(title, fontsize=16, fontweight='bold')
    for supplier, y in y_dict.items():
        plt.plot(x[supplier], y, label=supplier, linewidth=2, marker='o', markersize=6, linestyle='--')
    plt.legend(loc='best', fontsize=12)
    plt.xlabel(xlabel, fontsize=14)
    plt.ylabel(ylabel, fontsize=14)
    plt.xticks(fontsize=8)
    plt.yticks(fontsize=8)
    plt.tight_layout()
    graph = get_graph()
    return graph

def get_box_plot(x, title):
    plt.switch_backend('AGG')
    plt.figure(figsize=(9,5))
    plt.title(title)
    plt.boxplot(x)
    plt.tight_layout()
    graph = get_graph()
    return graph

def get_bar(x, y, title, ylabel, xlabel):
    plt.switch_backend('AGG')
    plt.figure(figsize=(9, 5))
    plt.title(title)
    plt.bar(x, y)
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)
    plt.tight_layout()
    graph = get_graph()
    return graph

def get_scatter(x, y, title, xlabel, ylabel):
    plt.switch_backend('AGG')
    plt.figure(figsize=(9,5))
    plt.title(title)
    plt.scatter(x, y)
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)
    plt.tight_layout()
    graph = get_graph()
    return graph